# descriptions, countries of origin, services, and contacts.
# I need an excel sheet with each category in the previous
# sentence as a column and the information extracted as the
#  rows (1 for each company).
from selenium import webdriver
# # from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl

driver = webdriver.Firefox()
driver.get(r"https://idexuae.ae/exhibitor/exhibitor-list/")
time.sleep(10)


wb = openpyxl.load_workbook(filename="idex.xlsx")
ws = wb.active
ind = 0


def eachPage():
    global ind
    for b in range(10):
        companyTitles = driver.find_elements_by_class_name('exhibitor-title')
        nameCell = f"A{10*(pages-1)+ind+1}"
        try:
            ws[nameCell] = companyTitles[b].text
        except IndexError:
            time.sleep(15)
            companyTitles = driver.find_elements_by_class_name('exhibitor-title')
            ws[nameCell] = companyTitles[b].text
        print("name:", companyTitles[b].text)
        neymar = companyTitles[b].text
        companyTitles[b].click()
        time.sleep(15)
        descCell = f"B{10*(pages-1)+ind}"
        contactsCell = f"C{10*(pages-1)+ind}"
        servicesCell = f"D{10*(pages-1)+ind}"
        try:
            # to get the descriptions
            descElement = driver.find_element_by_xpath('//div[@class="exhibitor-description"]//div[@class="holder"]//div//div')
            ws[descCell] = descElement.text
        except NoSuchElementException:
            ws[descCell] = ''
        try:
            # to get the contacts
            contacts = ''
            contactsElements = driver.find_elements_by_xpath('//div[@class="page-contact-details"]//div')
            for x in contactsElements:
                contacts += x.text + " "
            ws[contactsCell] = contacts
        except NoSuchElementException:
            ws[contactsCell] = ''
        try:
            # to get the country
            servicesElements = driver.find_elements_by_xpath('//div[@class="page-contact-details"]//div')
            services = ''
            for x in servicesElements:
                services += x.text + " "
            ws[servicesCell] = services
        except NoSuchElementException:
            ws[servicesCell] = ''
        driver.find_element_by_class_name('ma-0.v-btn').click()
        time.sleep(15)
        try:
            driver.find_element_by_class_name('ma-0.v-btn').click()
            time.sleep(15)
        except:
            pass
        print("done with", neymar)
        ind += 1
    wb.save("idex.xlsx")


for pages in range(3, 85):
    # find the company names and store them for each page
    # companyNames = [x.text for x in companyTitles]
    print('got companyNames', pages)
    pagenumbs = driver.find_elements_by_xpath('//ul//li//button')
    stopped = 31
    for page in pagenumbs:
        if page.text == str(pages):
            pagenum = page.text
            page.click()
            time.sleep(25)
            if int(pagenum) >= stopped:
                eachPage()
            break


wb.save("idex.xlsx")
